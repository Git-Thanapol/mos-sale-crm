"""Write path for /products: create/merge, bulk activate/deactivate,
archive/restore, and Excel import. There is no hard-delete path anywhere
in this app (invariant 8) — crm.catalog.selectors.product_delete_readiness
is a report only, never a gate feeding an actual delete.
"""

from __future__ import annotations

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from crm.catalog.models import Product
from crm.core.identity import clean

DEFAULT_PRODUCT_GROUP = "ทั่วไป"


class WorkbookFormatError(Exception):
    pass


@transaction.atomic
def create_or_merge_product(*, sku: str, product_name: str, product_group: str, actor_email: str) -> Product:
    """Matches legacy upsert_product_options: an exact (sku, product_group,
    product_name) match merges into the existing row (reactivates it and
    bumps the audit fields) instead of erroring on the unique constraint.
    """
    sku = clean(sku)
    product_name = clean(product_name)
    product_group = clean(product_group) or DEFAULT_PRODUCT_GROUP

    existing = Product.objects.filter(sku=sku, product_group=product_group, product_name=product_name).first()
    if existing:
        existing.is_active = True
        existing.updated_by = actor_email
        existing.save(update_fields=["is_active", "updated_by", "updated_at"])
        return existing

    return Product.objects.create(
        sku=sku,
        product_name=product_name,
        product_group=product_group,
        is_active=True,
        created_by=actor_email,
        updated_by=actor_email,
    )


def bulk_set_active(product_ids: list[int], is_active: bool, actor_email: str) -> int:
    return Product.objects.filter(id__in=product_ids).update(
        is_active=is_active, updated_by=actor_email, updated_at=timezone.now()
    )


def archive_products(product_ids: list[int], reason: str, actor_email: str) -> dict:
    reason = clean(reason) or "Archived from Product Master"
    now = timezone.now()
    updated = Product.objects.filter(id__in=product_ids, archived_at__isnull=True).update(
        archived_at=now, archived_by=actor_email, archive_reason=reason, is_active=False,
        updated_by=actor_email, updated_at=now,
    )
    return {"requested": len(product_ids), "updated": updated, "skipped": len(product_ids) - updated}


def restore_products(product_ids: list[int], actor_email: str) -> dict:
    # Restored products always come back deactivated — never auto-reactivated.
    now = timezone.now()
    updated = Product.objects.filter(id__in=product_ids, archived_at__isnull=False).update(
        archived_at=None, archived_by="", archive_reason="", is_active=False,
        updated_by=actor_email, updated_at=now,
    )
    return {"requested": len(product_ids), "updated": updated, "skipped": len(product_ids) - updated}


def _is_header_row(row) -> bool:
    cells = [clean(c).lower() for c in (row[:3] if row else [])]
    return cells == ["sku", "ชื่อสินค้า", "กลุ่มสินค้า"]


def import_products_workbook(file_obj, actor_email: str) -> dict:
    """Column A=SKU, B=ชื่อสินค้า, C=กลุ่มสินค้า, read positionally; a header
    row matching those exact labels is auto-detected and skipped. New-only:
    rows matching an existing DB product (or duplicated within the file) on
    the exact (sku, product_group, product_name) triple are skipped, not
    merged — a plain bulk insert, matching legacy's "confirm import"
    (which uses insert_product_options, not the create form's upsert).
    """
    try:
        workbook = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises several distinct types for a bad file
        raise WorkbookFormatError(f"ไม่สามารถเปิดไฟล์ได้: {exc}") from exc

    sheet = workbook.worksheets[0]
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise WorkbookFormatError("ไฟล์ไม่มีข้อมูล")
    if len(all_rows[0]) < 3:
        raise WorkbookFormatError("ไฟล์ต้องมีอย่างน้อย 3 คอลัมน์: SKU / ชื่อสินค้า / กลุ่มสินค้า")

    data_rows = all_rows[1:] if _is_header_row(all_rows[0]) else all_rows

    existing_keys = set(Product.objects.values_list("sku", "product_group", "product_name"))
    seen_in_file: set[tuple[str, str, str]] = set()
    created = duplicate = invalid = 0

    with transaction.atomic():
        for row in data_rows:
            sku = clean(row[0]) if len(row) > 0 else ""
            product_name = clean(row[1]) if len(row) > 1 else ""
            product_group = clean(row[2]) if len(row) > 2 else ""

            if not sku or not product_name or not product_group:
                invalid += 1
                continue

            key = (sku, product_group, product_name)
            if key in existing_keys or key in seen_in_file:
                duplicate += 1
                continue

            seen_in_file.add(key)
            Product.objects.create(
                sku=sku, product_name=product_name, product_group=product_group,
                is_active=True, created_by=actor_email, updated_by=actor_email,
            )
            created += 1

    return {"created": created, "duplicate": duplicate, "invalid": invalid}
