"""The shared Excel-import implementation — used by both
`manage.py import_xlsx` (Phase 2, CLI) and the /orders/import web view
(Phase 4). One implementation, two entry points, so they can't drift.

Two fixed templates, auto-detected from the header row, no column-mapping
UI: the classic one-line-per-order-line template (see the module docstring
in the import_xlsx management command for the full header list and
rationale) and the wide multi-SKU template (see is_wide_format /
iter_wide_records below). Writes to BOTH the staging table and the
normalized core in one transaction per row, applying the multi-SKU merge
rule (invariant 7).
"""

from __future__ import annotations

import uuid
from datetime import date
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from crm.core.identity import clean, make_dedupe_key, normalize_phone, parse_date, phone_key
from crm.customers.models import Customer
from crm.imports.models import StagingImportRow
from crm.online_orders.importers import is_online_detail_format, is_online_main_format
from crm.online_orders.services import import_online_detail, import_online_main
from crm.orders.models import Order, OrderLine

HEADER_MAP = {
    "เลขคำสั่งซื้อ": "order_id",
    "ชื่อลูกค้า": "customer_name",
    "เบอร์โทร": "phone1",
    "เบอร์สำรอง": "phone2",
    "SKU": "sku",
    "ชื่อสินค้า": "product_name",
    "จำนวน": "quantity",
    "ราคา": "amount",
    "วันที่สั่งซื้อ": "order_date",
    "ประเภทการขาย": "sale_type",
    "จังหวัด": "province",
    "อำเภอ": "city",
    "รหัสไปรษณีย์": "postal_code",
    "ที่อยู่": "address",
    "ผู้ดูแล": "owner",
    "รหัสพนักงาน": "staff_code",
    "เลขพัสดุ": "tracking_no",
    "ขนส่ง": "carrier",
    "สถานะ": "order_status",
}
REQUIRED_FIELDS = ("customer_name",)

EMPTY_COUNTS = {"valid": 0, "invalid": 0, "customers_created": 0, "orders_created": 0, "lines_created": 0}

# The wide multi-SKU export template: one row per order, up to 6 SKU/qty/price
# column-sets, date split into วันที่/เดือน/ปี, no staff_code column. Detected
# by the presence of "SKU (1)" in the header, since it shares no column names
# with the classic fixed template above. See iter_wide_records for the
# reshaping rule — originally established (and confirmed with the customer,
# 2026-07-26) for the one-off crm.imports.management.commands
# .import_legacy_wide_xlsx historical import; reused here as the shape for
# ongoing uploads of files in this format.
WIDE_MARKER_HEADER = "SKU (1)"
WIDE_LINE_SLOTS = range(1, 7)
WIDE_OPENING_STAFF_KEYS = ("พนักงานขาย", "พนักงานเปิดบิล")


class WorkbookFormatError(Exception):
    pass


def to_decimal(value) -> Decimal | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def resolve_columns(header_row) -> dict[str, int]:
    columns = {}
    for idx, header in enumerate(header_row):
        field = HEADER_MAP.get(clean(header))
        if field:
            columns[field] = idx

    missing_required = [h for h, f in HEADER_MAP.items() if f in REQUIRED_FIELDS and f not in columns]
    if missing_required:
        raise WorkbookFormatError(f"missing required column(s): {missing_required}")
    if "phone1" not in columns and "phone2" not in columns:
        raise WorkbookFormatError("workbook must have at least one of เบอร์โทร / เบอร์สำรอง")
    return columns


def is_wide_format(header_row) -> bool:
    return any(clean(h) == WIDE_MARKER_HEADER for h in header_row)


def _wide_id_like(value) -> str:
    """For cells that are sometimes floats (postal code, tracking number)
    because Excel stored a numeric-looking string as a number — strip the
    trailing '.0' a plain clean(value) would otherwise leave in.
    """
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def _wide_order_date(day, month, year) -> str:
    try:
        d, m, y = int(day), int(month), int(year)
        return f"{y:04d}-{m:02d}-{d:02d}"
    except (TypeError, ValueError):
        return ""


def iter_wide_records(header_row, data_rows):
    """Reshapes each wide-template row into 1-6 flat per-line-item records
    in import_row()'s shape (one dict per populated SKU line). Business
    rules (see module comment above WIDE_MARKER_HEADER):

    - No per-SKU price on a row -> the row's ยอดขายรวม total is assigned to
      the first populated SKU line only; other lines get amount=None.
    - owner = พนักงานดูแล (care) if populated, else พนักงานขาย /
      พนักงานเปิดบิล (opening/sales staff — column name varies by export).
      staff_code is always blank — the template has no staff_code column.
    - sale_type is always NEW_ORDER — the template has no sale_type column.
    - Rows with a blank เลขคำสั่งซื้อ, or with no populated SKU/product name
      across all 6 slots, are skipped.

    Yields (record, row_number) pairs.
    """
    idx = {clean(h): i for i, h in enumerate(header_row)}

    def get(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    opening_key = next((k for k in WIDE_OPENING_STAFF_KEYS if k in idx), None)
    row_number = 1

    for row in data_rows:
        row_number += 1
        order_id = clean(get(row, "เลขคำสั่งซื้อ"))
        if not order_id:
            continue

        care = clean(get(row, "พนักงานดูแล"))
        opening = clean(get(row, opening_key)) if opening_key else ""

        base_fields = {
            "order_id": order_id,
            "customer_name": clean(get(row, "ลูกค้า")),
            "phone1": clean(get(row, "เบอร์โทร (1)")),
            "phone2": clean(get(row, "เบอร์โทร (2)")),
            "order_date": _wide_order_date(get(row, "วันที่"), get(row, "เดือน"), get(row, "ปี")),
            "sale_type": "NEW_ORDER",
            "province": clean(get(row, "จังหวัด")),
            "city": clean(get(row, "อำเภอ")),
            "subdistrict": clean(get(row, "ตำบล")),
            "postal_code": _wide_id_like(get(row, "รหัสไปรษณีย์")),
            "address": clean(get(row, "ที่อยู่จัดส่ง")),
            "owner": care or opening,
            "staff_code": "",
            "tracking_no": _wide_id_like(get(row, "หมายเลขพัสดุ")),
            "carrier": clean(get(row, "ขนส่ง")),
            "order_status": clean(get(row, "สถานะคำสั่งซื้อ")),
        }

        lines = []
        for i in WIDE_LINE_SLOTS:
            sku = clean(get(row, f"SKU ({i})"))
            product_name = clean(get(row, f"สินค้า ({i})"))
            if not sku and not product_name:
                continue
            lines.append({
                "sku": sku,
                "product_name": product_name,
                "quantity": get(row, f"จำนวน ({i})"),
                "amount": get(row, f"ราคา ({i})"),
            })
        if not lines:
            continue

        if all(ln["amount"] in (None, "") for ln in lines):
            lines[0]["amount"] = get(row, "ยอดขายรวม")

        for line in lines:
            yield {**base_fields, **line}, row_number


@transaction.atomic
def import_row(record: dict, row_number: int, batch_id, uploaded_by: str) -> dict:
    customer_name = clean(record.get("customer_name"))
    phone1 = normalize_phone(record.get("phone1"))
    phone2 = normalize_phone(record.get("phone2"))
    validation_error = ""
    if not customer_name:
        validation_error = "ชื่อลูกค้า is required"
    elif not phone1 and not phone2:
        validation_error = "requires at least one of เบอร์โทร / เบอร์สำรอง"

    import_status = "invalid" if validation_error else "valid"
    quantity = to_decimal(record.get("quantity"))
    amount = to_decimal(record.get("amount"))
    sale_type = clean(record.get("sale_type")) or "NEW_ORDER"
    # parse_date() returns an ISO string by contract (see crm.core.identity,
    # matches the pinned legacy test), so convert to a real date object
    # here for the model fields — assigning a string straight into a
    # DateField leaves the in-memory attribute a str until reloaded from
    # the DB, breaking date comparisons below.
    order_date_str = parse_date(record.get("order_date"))
    order_date = date.fromisoformat(order_date_str) if order_date_str else None

    staging = StagingImportRow.objects.create(
        import_batch_id=batch_id,
        row_number=row_number,
        uploaded_by=uploaded_by,
        raw_data={k: ("" if v is None else str(v)) for k, v in record.items()},
        order_id=clean(record.get("order_id")),
        customer_name=customer_name,
        phone1=phone1,
        phone2=phone2,
        sku=clean(record.get("sku")),
        product_name=clean(record.get("product_name")),
        quantity=quantity,
        order_date=order_date,
        province=clean(record.get("province")),
        city=clean(record.get("city")),
        subdistrict=clean(record.get("subdistrict")),
        postal_code=clean(record.get("postal_code")),
        address=clean(record.get("address")),
        owner=clean(record.get("owner")),
        staff_code=clean(record.get("staff_code")),
        tracking_no=clean(record.get("tracking_no")),
        carrier=clean(record.get("carrier")),
        order_status=clean(record.get("order_status")),
        sale_type=sale_type,
        amount=amount,
        total_amount=amount,
        import_status=import_status,
        validation_error=validation_error,
        dedupe_key=make_dedupe_key(record.get("order_id"), phone1, phone2, record.get("tracking_no")),
    )

    if import_status == "invalid":
        return {**EMPTY_COUNTS, "invalid": 1}

    result = dict(EMPTY_COUNTS, valid=1)
    key = phone_key(phone1, phone2, staging.pk)
    customer, created = Customer.objects.get_or_create(
        phone_key=key,
        defaults={
            "phone1": phone1,
            "phone2": phone2,
            "customer_name": customer_name,
            "province": staging.province,
            "city": staging.city,
            "subdistrict": staging.subdistrict,
            "postal_code": staging.postal_code,
            "address": staging.address,
            "owner_display": staging.owner,
            "staff_code": staging.staff_code,
            "source_import_row": staging,
        },
    )
    if created:
        result["customers_created"] = 1

    order = None
    order_no = staging.order_id
    if order_no:
        order = Order.objects.filter(customer=customer, order_no=order_no).first()
    if order is None:
        order = Order.objects.create(
            customer=customer,
            order_no=order_no,
            order_date=staging.order_date,
            sale_type=sale_type,
            order_status=staging.order_status,
            carrier=staging.carrier,
            tracking_no=staging.tracking_no,
            total_amount=amount,  # placeholder: this row's line amount; recompute_rollups sums all lines
            owner_display=staging.owner,
            staff_code=staging.staff_code,
            source_type="import",
            uploaded_by=uploaded_by,
            import_batch_id=batch_id,
            source_import_row=staging,
        )
        result["orders_created"] = 1

    if staging.sku or staging.product_name:
        line, line_created = OrderLine.objects.get_or_create(
            order=order,
            sku=staging.sku,
            product_name=staging.product_name,
            defaults={"quantity": quantity or 1, "amount": amount, "source_import_row": staging},
        )
        if not line_created and quantity:
            line.quantity += int(quantity)
            line.save(update_fields=["quantity"])
        elif line_created:
            result["lines_created"] = 1

    if customer.last_order_date is None or (staging.order_date and staging.order_date >= customer.last_order_date):
        customer.last_order_date = staging.order_date
        customer.last_order = order
    customer.order_count = customer.orders.count()
    customer.save(update_fields=["last_order_date", "last_order", "order_count", "updated_at"])

    return result


def import_workbook(file_obj, uploaded_by: str) -> dict:
    """The single entry point: parses + validates + commits every row.
    Raises WorkbookFormatError before anything is written if the header
    row is unusable. Supports four header shapes, auto-detected from the
    header row: the two online-order platform exports (see
    crm.online_orders.importers, checked first since they're identified by
    signature, not by resolve_columns' required-column probe), the classic
    fixed template (one row per order line), and the wide multi-SKU
    template (see is_wide_format/iter_wide_records). Returns the same
    counts dict shape the CLI prints either way.
    """
    try:
        workbook = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises several distinct types for a bad file
        raise WorkbookFormatError(f"ไม่สามารถเปิดไฟล์ได้: {exc}") from exc

    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise WorkbookFormatError("ไฟล์ไม่มีข้อมูล")

    header_row, data_rows = rows[0], rows[1:]
    batch_id = uuid.uuid4()
    counts = dict(EMPTY_COUNTS)

    if is_online_main_format(header_row):
        return import_online_main(data_rows, batch_id, uploaded_by)
    if is_online_detail_format(header_row):
        return import_online_detail(data_rows, batch_id, uploaded_by)

    if is_wide_format(header_row):
        records = iter_wide_records(header_row, data_rows)
    else:
        columns = resolve_columns(header_row)
        records = (
            ({field: row[idx] if idx < len(row) else None for field, idx in columns.items()}, row_number)
            for row_number, row in enumerate(data_rows, start=2)
        )

    for record, row_number in records:
        row_result = import_row(record, row_number, batch_id, uploaded_by)
        for key, value in row_result.items():
            counts[key] += value

    return {"batch_id": batch_id, **counts}
