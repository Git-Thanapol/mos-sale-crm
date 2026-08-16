"""One-off importer for the real historical wide-format export
(`DATA <year>.xlsx`: one row per order, up to 6 SKU/qty/price
column-sets, date split into วันที่/เดือน/ปี, staff recorded as
free-text names across up to 3 role columns, no staff_code column, no
sale_type column). This is NOT the same shape as the standard
fixed-template importer (crm.imports.services.import_workbook) — that
one already expects one row per order line in the app's canonical
column names. This command reshapes each wide row into 1-6 flat
per-line-item records in that same shape and calls
crm.imports.services.import_row() for each one, reusing its proven
customer/order/line find-or-create + rollup logic rather than
reimplementing it.

Business-logic choices made here (confirmed with the customer
2026-07-26, see conversation — no ticket/doc reference exists since
this was a one-off decision, not a standing policy):

- No per-SKU price on a row -> the full order-level ยอดขายรวม total is
  assigned to the row's FIRST populated SKU line only; any other lines
  on that same order get amount=None. Affects the 2568 file, which has
  zero populated ราคา (n) cells across its ~73k real rows.
- owner_display = พนักงานดูแล (care) if populated, else พนักงานขาย /
  พนักงานเปิดบิล (opening/sales staff — the column name varies by
  year). staff_code is left blank for every row: the source file has
  no staff_code column at all, so these orders are invisible to any
  staff_code-scoped telesell login until someone assigns staff_code by
  hand later. That's crm.core.scoping's normal fail-closed behavior on
  a blank staff_code, not a bug introduced here.
- sale_type is always NEW_ORDER. The source has no sale_type column,
  and พนักงานอัพเซลล์ (upsell staff) presence (~5% of rows) was
  deliberately NOT used to infer UPSELL — ambiguous whether it means
  "this whole order is an upsell" vs "an upsell agent also touched
  this order", and getting it wrong would misclassify revenue.
- Rows with a blank เลขคำสั่งซื้อ (order number) are skipped — these are
  trailing blank rows from the sheet's used-range extending past the
  real data (openpyxl still yields them via iter_rows), not real orders.
"""

from __future__ import annotations

import uuid

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from crm.imports.services import EMPTY_COUNTS, import_row

OPENING_STAFF_KEYS = ("พนักงานขาย", "พนักงานเปิดบิล")
ZERO_WIDTH_SPACE = "​"


def _get(row, idx, key):
    i = idx.get(key)
    return row[i] if i is not None and i < len(row) else None


def _clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace(ZERO_WIDTH_SPACE, "").strip()


def _clean_id_like(value) -> str:
    """For cells that are sometimes floats (postal code, tracking number)
    because Excel stored a numeric-looking string as a number — strip the
    trailing '.0' a plain str() would otherwise leave in.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_text(value)


def _reconstruct_date(day, month, year) -> str:
    try:
        d, m, y = int(day), int(month), int(year)
        return f"{y:04d}-{m:02d}-{d:02d}"
    except (TypeError, ValueError):
        return ""


class Command(BaseCommand):
    help = "One-off importer for the real historical wide-format .xlsx export (see module docstring)."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the .xlsx file")
        parser.add_argument(
            "--uploaded-by", default="",
            help="Value stored in Order.uploaded_by — usually left blank for historical data (not an email)",
        )

    def handle(self, *args, **options):
        path = options["path"]
        uploaded_by = options["uploaded_by"]
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises several distinct types for a bad file
            raise CommandError(f"cannot open {path}: {exc}") from exc

        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError("workbook has no header row")
        idx = {h: i for i, h in enumerate(header)}

        opening_key = next((k for k in OPENING_STAFF_KEYS if k in idx), None)

        batch_id = uuid.uuid4()
        counts = dict(EMPTY_COUNTS)
        row_number = 1
        orders_seen = 0

        for row in rows:
            row_number += 1
            order_id = _clean_text(_get(row, idx, "เลขคำสั่งซื้อ"))
            if not order_id:
                continue
            orders_seen += 1

            order_date = _reconstruct_date(
                _get(row, idx, "วันที่"), _get(row, idx, "เดือน"), _get(row, idx, "ปี")
            )
            care = _clean_text(_get(row, idx, "พนักงานดูแล"))
            opening = _clean_text(_get(row, idx, opening_key)) if opening_key else ""

            base_fields = {
                "order_id": order_id,
                "customer_name": _clean_text(_get(row, idx, "ลูกค้า")),
                "phone1": _clean_text(_get(row, idx, "เบอร์โทร (1)")),
                "phone2": _clean_text(_get(row, idx, "เบอร์โทร (2)")),
                "order_date": order_date,
                "sale_type": "NEW_ORDER",
                "province": _clean_text(_get(row, idx, "จังหวัด")),
                "city": _clean_text(_get(row, idx, "อำเภอ")),
                "postal_code": _clean_id_like(_get(row, idx, "รหัสไปรษณีย์")),
                "address": _clean_text(_get(row, idx, "ที่อยู่จัดส่ง")),
                "owner": care or opening,
                "staff_code": "",
                "tracking_no": _clean_id_like(_get(row, idx, "หมายเลขพัสดุ")),
                "carrier": _clean_text(_get(row, idx, "ขนส่ง")),
                "order_status": _clean_text(_get(row, idx, "สถานะคำสั่งซื้อ")),
            }

            lines = []
            for i in range(1, 7):
                sku = _clean_text(_get(row, idx, f"SKU ({i})"))
                product_name = _clean_text(_get(row, idx, f"สินค้า ({i})"))
                if not sku and not product_name:
                    continue
                lines.append({
                    "sku": sku,
                    "product_name": product_name,
                    "quantity": _get(row, idx, f"จำนวน ({i})"),
                    "amount": _get(row, idx, f"ราคา ({i})"),
                })
            if not lines:
                continue

            # No per-line price anywhere on this row -> put the order-level
            # total on the first line only (see module docstring).
            if all(ln["amount"] in (None, "") for ln in lines):
                lines[0]["amount"] = _get(row, idx, "ยอดขายรวม")

            for line in lines:
                record = {**base_fields, **line}
                result = import_row(record, row_number, batch_id, uploaded_by)
                for key, value in result.items():
                    counts[key] += value

        self.stdout.write(self.style.SUCCESS(f"orders_seen={orders_seen} batch={batch_id} {counts}"))
