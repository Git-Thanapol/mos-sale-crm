from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from crm.accounts.models import User
from crm.customers.export import CRM_EXPORT_HEADERS
from crm.customers.models import Customer
from crm.orders.models import Order, OrderLine

pytestmark = [pytest.mark.view, pytest.mark.django_db]


def test_export_requires_can_export_customers(client):
    telesell = User.objects.create_user(email="staff@example.com", password="x", role="พนักงาน", staff_code="S0001")
    telesell.must_change_password = False
    telesell.save()
    client.force_login(telesell)

    resp = client.get(reverse("customers:export"))
    assert resp.status_code == 403


def test_admin_cannot_export_either(client):
    """ADMIN is weaker than EDITOR here too — preserved verbatim."""
    admin = User.objects.create_user(email="admin@example.com", password="x", role="ADMIN")
    admin.must_change_password = False
    admin.save()
    client.force_login(admin)

    resp = client.get(reverse("customers:export"))
    assert resp.status_code == 403


def test_export_headers_match_template(client):
    editor = User.objects.create_user(email="editor@example.com", password="x", role="EDITOR")
    editor.must_change_password = False
    editor.save()
    client.force_login(editor)

    resp = client.get(reverse("customers:export"))
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    assert header_row == CRM_EXPORT_HEADERS


def test_export_one_row_per_order_line(client):
    editor = User.objects.create_user(email="editor2@example.com", password="x", role="EDITOR")
    editor.must_change_password = False
    editor.save()
    client.force_login(editor)

    customer = Customer.objects.create(
        phone_key="0899600001", phone1="0899600001", customer_name="Export Test", url="https://x.test"
    )
    order = Order.objects.create(customer=customer, order_no="EXPORD1", order_date="2026-07-01")
    OrderLine.objects.create(order=order, sku="SP001", product_name="A", quantity=1, amount=100)
    OrderLine.objects.create(order=order, sku="SP002", product_name="B", quantity=2, amount=200)

    resp = client.get(reverse("customers:export"), {"keyword": "Export Test"})
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    data_rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    matching = [r for r in data_rows if r[1] == "EXPORD1"]
    assert len(matching) == 2  # one row per line, not one per order
    skus = {r[3] for r in matching}
    assert skus == {"SP001", "SP002"}
